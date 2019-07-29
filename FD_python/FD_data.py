import os
from scipy import stats
import numpy as np
import openpyxl
from matplotlib import pyplot as plt

def FD_merge_dataloader():
    def normalize_col(X_, axis=0):
        print("normalize step")
        assert len(np.amax(X_, axis=axis)) == len(X_[0])
        assert np.all(np.amax(X_, axis=axis) != 0)
        X_ = X_ - np.amin(X_, axis=axis)
        return (X_ - np.amin(X_, axis=axis)) / np.amax(X_, axis=axis)

    # ----------------------- excel data read -------------------#
    base_folder_path = '/home/soopil/Desktop/Dataset/brain_ewha'  # desktop setting
    train_path = os.path.join(base_folder_path, 'Train_Meningioma_20180508.xlsx')
    test_path = os.path.join(base_folder_path, 'Test_Meningioma_20180508.xlsx')

    xl = openpyxl.load_workbook(train_path, read_only=True)
    ws = xl['20171108_New_N4ITK corrected']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)

    feature_num = 102 - 4
    train_x = data_excel[1:, 4:4+feature_num]
    train_y = data_excel[1:, 2]
    train_subj = data_excel[1:,0]
    train_x = np.delete(train_x, 109, 0)
    train_y = np.delete(train_y, 109, 0)
    train_subj = np.delete(train_subj, 109, 0)

    xl = openpyxl.load_workbook(test_path, read_only=True)
    ws = xl['Sheet1']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)

    test_x = data_excel[1:, 4:4+feature_num]
    test_y = data_excel[1:, 2]
    test_subj = data_excel[1:, 0]
    test_x = np.delete(test_x, 59, 0)
    test_y = np.delete(test_y, 59, 0)
    test_subj = np.delete(test_subj, 59, 0)

    is_norm = True
    if is_norm:
        train_x = normalize_col(train_x, axis=0)
        test_x = normalize_col(test_x, axis=0)

    # ----------------------- FD data read -------------------#

    result_file_name = './fd_result/fd_result_file.txt'
    fd = open(result_file_name, 'r')
    contents = fd.readlines()[1:]
    # print(contents)

    fd_list, name_list = [], []
    for e in contents:
        subj_name = e.split('/')[0]
        fd = e.split('/')[1].replace('\n', '').split(',')
        if subj_name in name_list:
            continue
        fd_list.append([subj_name, fd])
        name_list.append(subj_name)

    tr_x, tr_y, tst_x, tst_y = [],[],[],[]
    valid_dim = 9
    cnt = 0
    for subj in fd_list:
        name = int(subj[0])
        n = np.array(subj[1]).astype(np.float32)
        length = len(n)
        r = np.array([2 ** i for i in range(0, length)]).astype(np.float32)
        n_grad = np.gradient(np.log(n))
        r_grad = np.gradient(np.log(r))
        dim = (np.divide(n_grad, r_grad) * (-1))[:valid_dim]
        # print(len(dim))
        # print(dim)
        # print(name, type(name))
        if name in train_subj:
            tr_index = np.where(train_subj == name)
            # print(tr_index, train_x[tr_index], train_y[tr_index])
            x = list(dim)
            # x = list(dim) + list(np.squeeze(train_x[tr_index]))
            y = train_y[tr_index][0]
            tr_x.append(x)
            tr_y.append(y)
            # assert False
        elif name in test_subj:
            tst_index = np.where(test_subj == name)
            x = list(dim)
            # x = list(dim) + list(np.squeeze(test_x[tst_index]))
            y = test_y[tst_index][0]
            tst_x.append(x)
            tst_y.append(y)
            pass
        else:
            cnt += 1

    tr_x, tst_x = np.array(tr_x), np.array(tst_x)
    is_norm = True
    if is_norm:
        tmp_x = np.concatenate([tr_x[:,:valid_dim], tst_x[:,:valid_dim]], axis=0)
        mini, maxi = np.amin(tmp_x[:,:valid_dim]), np.amax(tmp_x[:,:valid_dim])
        tr_x[:,:valid_dim] = (tr_x[:,:valid_dim] - mini) / maxi
        tst_x[:,:valid_dim] = (tst_x[:,:valid_dim] - mini) / maxi
        # tr_x[:valid_dim] = normalize_(tr_x)
        # tst_x = normalize_col(tst_x, axis=0)
    # tr_x, tr_y, tst_x, tst_y = np.array(tr_x), np.array(tr_y), np.array(tst_x), np.array(tst_y)

    print(np.shape(tr_x), np.shape(tr_y), np.shape(tst_x), np.shape(tst_y))
    # print(train_x.shape, train_y.shape, test_x.shape, test_y.shape)
    return tr_x, tr_y, tst_x, tst_y

def FD_dataloader(xl_path, sheet_name, result_file_path='../fd_result/fd_result_file.txt'):
    xl = openpyxl.load_workbook(xl_path, read_only=True)
    ws = xl[sheet_name]
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)
    print(data_excel.shape)
    low_subj_pos = np.where(data_excel[:, 2] == 0)
    high_subj_pos = np.where(data_excel[:, 2] == 1)
    low_fdim, high_fdim = [], []
    low_subj_list = [data_excel[i, 0] for i in low_subj_pos][0]
    high_subj_list = [data_excel[i, 0] for i in high_subj_pos][0]
    print(np.shape(low_subj_list))
    print(np.shape(high_subj_list))
    print(type(high_subj_list))
    # result_file_path = '../fd_result/fd_result_file.txt'
    # result_file_path = '../fd_result/SINCHON_FD_result_20190723.txt'
    fd = open(result_file_path, 'r')
    contents = fd.readlines()[1:]
    # print(contents)

    fd_list, name_list = [], []
    for e in sorted(contents):
        subj_name = e.split('/')[0]
        fd = e.split('/')[1].replace('\n', '').split(',')
        # do not include subject if there is same subject name already
        if subj_name in name_list:
            continue
        fd_list.append([subj_name, fd])
        name_list.append(subj_name)
        if subj_name == '1550930':
            print(contents.index(e))

    l_c, h_c, max_c = 0, 0, 200
    high_subj_list_f, low_subj_list_f = [], []

    for i, e in enumerate(fd_list):
        sample = e
        subj_name = int(e[0])
        # print(subj_name)
        n = np.array(sample[1]).astype(np.float32)
        length = len(n)
        r = np.array([2 ** i for i in range(0, length)]).astype(np.float32)
        n_grad = np.gradient(np.log(n))
        r_grad = np.gradient(np.log(r))
        dim = np.divide(n_grad, r_grad) * (-1)
        # print(len(dim), end='/')
        if subj_name in low_subj_list and l_c < max_c:
            # print('low list')
            # low_fdim.append(length)
            low_fdim.append(dim[:9])
            # plt.plot(np.log(r), dim, '-bo')
            l_c += 1
            low_subj_list_f.append(subj_name)

        elif subj_name in high_subj_list and h_c < max_c:
            # print('high list')
            # high_fdim.append(length)
            high_fdim.append(dim[:9])
            # plt.plot(np.log(r), dim, '-ro')
            h_c += 1
            high_subj_list_f.append(subj_name)

    # print(high_subj_list_f)
    # print(low_subj_list_f)
    # assert False
    label_high = [1 for _ in range(h_c)]
    label_low = [0 for _ in range(l_c)]
    print()
    print(h_c,l_c) # 30 and 95
    # print(label_high)
    # print(label_low)
    return high_fdim, low_fdim, label_high, label_low, high_subj_list_f, low_subj_list_f

def FD_dataloader_save(xl_path, sheet_name):
    xl = openpyxl.load_workbook(xl_path, read_only=True)
    ws = xl[sheet_name]
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)
    print(data_excel.shape)
    low_subj_pos = np.where(data_excel[:, 2] == 0)
    high_subj_pos = np.where(data_excel[:, 2] == 1)
    low_fdim, high_fdim = [], []
    low_subj_list = [data_excel[i, 0] for i in low_subj_pos][0]
    high_subj_list = [data_excel[i, 0] for i in high_subj_pos][0]
    print(np.shape(low_subj_list))
    print(np.shape(high_subj_list))
    print(type(high_subj_list))
    result_file_name = './fd_result/fd_result_file.txt'
    fd = open(result_file_name, 'r')
    contents = fd.readlines()[1:]
    # print(contents)

    fd_list, name_list = [], []
    for e in contents:
        subj_name = e.split('/')[0]
        fd = e.split('/')[1].replace('\n', '').split(',')
        if subj_name in name_list:
            continue
        fd_list.append([subj_name, fd])
        name_list.append(subj_name)
        if subj_name == '1550930':
            print(contents.index(e))

    l_c, h_c, max_c = 0, 0, 200
    for i, e in enumerate(fd_list):
        sample = e
        subj_name = int(e[0])
        n = np.array(sample[1]).astype(np.float32)
        length = len(n)
        r = np.array([2 ** i for i in range(0, length)]).astype(np.float32)
        n_grad = np.gradient(np.log(n))
        r_grad = np.gradient(np.log(r))
        dim = np.divide(n_grad, r_grad) * (-1)
        print(len(dim), end='/')
        if subj_name in low_subj_list and l_c < max_c:
            print('low list')
            # low_fdim.append(length)
            low_fdim.append(dim[:9])
            plt.plot(np.log(r), dim, '-bo')
            l_c += 1
        elif subj_name in high_subj_list and h_c < max_c:
            print('high list')
            # high_fdim.append(length)
            high_fdim.append(dim[:9])
            plt.plot(np.log(r), dim, '-ro')
            h_c += 1

    label_high = [1 for _ in range(h_c)]
    label_low = [0 for _ in range(l_c)]
    print()
    print(h_c,l_c) # 30 and 95
    # print(label_high)
    # print(label_low)
    return high_fdim, low_fdim, label_high, label_low

def Lac_dataloader_save():
    xl_test = '/home/soopil/Desktop/Dataset/brain_ewha/Test_Meningioma_20180508.xlsx'
    xl_train = '/home/soopil/Desktop/Dataset/brain_ewha/Train_Meningioma_20180508.xlsx'
    xl = openpyxl.load_workbook(xl_train, read_only=True)
    ws = xl['20171108_New_N4ITK corrected']  # 'Sheet1 for test excel file
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)
    print(data_excel.shape)
    low_subj_pos = np.where(data_excel[:, 2] == 0)
    high_subj_pos = np.where(data_excel[:, 2] == 1)
    low_fdim, high_fdim = [], []
    low_subj_list = [data_excel[i, 0] for i in low_subj_pos][0]
    high_subj_list = [data_excel[i, 0] for i in high_subj_pos][0]
    print(np.shape(low_subj_list))
    print(np.shape(high_subj_list))
    print(type(high_subj_list))
    result_file_name = './fd_result/Lac_result_v2.txt' # Lac_result_v2.txt
    fd = open(result_file_name, 'r')
    contents = fd.readlines()[1:]
    # print(contents)
    fd_list = []
    for e in contents:
        subj_name = e.split('/')[0]
        fd = e.split('/')[1].replace('\n', '').split(',')
        if subj_name == '6651225':
            continue
        fd_list.append([subj_name, fd])

    for e in fd_list:
        print(e)
    # assert False
    fd_list = np.array(fd_list)
    l_c, h_c, max_c = 0, 0, 20
    for i, e in enumerate(fd_list):
        sample = e
        subj_name = int(e[0])
        n = np.array(sample[1])[:5].astype(np.float32)
        # n = np.ones_like(n) / n
        # n = np.log(n)
        length = len(n)
        r = np.array([2 ** i for i in range(0, length)]).astype(np.float32)
        # print(n)
        for k in n:
            if str(k)=='nan':
                print(subj_name, n,e)
                assert False
            print(k, str(k)=='nan', end='/')
        print()
        # print(np.where(str(n) == 'nan'))
        # n_grad = np.gradient(np.log(n))
        # r_grad = np.gradient(np.log(r))
        if subj_name in low_subj_list and l_c < max_c:
            print('low list')
            # low_fdim.append(length)
            low_fdim.append(n)
            plt.plot(np.log(r), n, '-bo')
            l_c += 1
        elif subj_name in high_subj_list and h_c < max_c:
            print('high list')
            # high_fdim.append(length)
            high_fdim.append(n)
            plt.plot(np.log(r), n, '-ro')
            h_c += 1
    label_high = [1 for _ in range(h_c)]
    label_low = [0 for _ in range(l_c)]
    plt.show()
    print()
    print(h_c,l_c) # 30 and 95
    # print(label_high)
    # print(label_low)
    tTestResult = stats.ttest_ind(low_fdim, high_fdim)
    print(tTestResult)
    return high_fdim+low_fdim, label_high+label_low

def Lac_dataloader(xl_path, sheet_name, result_file_path = '../fd_result/Lac_result_v2.txt'):
    xl = openpyxl.load_workbook(xl_path, read_only=True)
    ws = xl[sheet_name]  # 'Sheet1 for test excel file
    low_fdim, high_fdim, data_excel = [], [] , []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)
    print(data_excel.shape)

    low_subj_pos = np.where(data_excel[:, 2] == 0)
    high_subj_pos = np.where(data_excel[:, 2] == 1)
    low_subj_list = [data_excel[i, 0] for i in low_subj_pos][0]
    high_subj_list = [data_excel[i, 0] for i in high_subj_pos][0]
    print(np.shape(low_subj_list))
    print(np.shape(high_subj_list))
    print(type(high_subj_list))

    # result_file_path = '../fd_result/Lac_result_v2.txt' # Lac_result_v2.txt
    fd = open(result_file_path, 'r')
    contents = fd.readlines()[1:]
    fd_list, name_list = [], []
    for e in contents:
        subj_name = e.split('/')[0]
        fd = e.split('/')[1].replace('\n', '').split(',')
        if subj_name == '6651225':
            continue
        if subj_name in name_list:
            continue
        fd_list.append([subj_name, fd])
        name_list.append(subj_name)

    # name_list = np.array(name_list)
    # print(name_list)
    # for n in name_list:
    #     index = np.where(name_list == n)
    #     print(index[0], len(index[0]))
    #
    #     # print(name_list[index])
    #     if len(index[0]) > 1:
    #         print(index)
    #
    #         for i in index[0]:
    #             print(fd_list[i])
    # assert False
    # for e in fd_list:
    #     print(e)
    # assert False
    fd_list = np.array(fd_list)
    l_c, h_c, max_c = 0, 0, 200
    high_subj_list_f, low_subj_list_f = [], []

    for i, e in enumerate(fd_list):
        sample = e
        subj_name = int(e[0])
        n = np.array(sample[1])[:5].astype(np.float32)
        # n = np.ones_like(n) / n
        n = np.log(n)

        for k in n:
            if str(k)=='nan':
                print(subj_name, n,e)
                assert False
            # print(k, str(k)=='nan', end='/')
        # print()

        if subj_name in low_subj_list and l_c < max_c:
            # print('low list')
            low_fdim.append(n)
            low_subj_list_f.append(subj_name)
            l_c += 1
        elif subj_name in high_subj_list and h_c < max_c:
            # print('high list')
            high_fdim.append(n)
            high_subj_list_f.append(subj_name)
            h_c += 1

    label_high = [1 for _ in range(h_c)]
    label_low = [0 for _ in range(l_c)]
    # plt.show()
    print()
    print(h_c,l_c)
    return high_fdim, low_fdim, label_high, label_low, high_subj_list_f, low_subj_list_f


if __name__ == "__main__":
    xl_test = '/home/soopil/Desktop/Dataset/brain_ewha_early/Test_Meningioma_20180508.xlsx'
    xl_train = '/home/soopil/Desktop/Dataset/brain_ewha_early/Train_Meningioma_20180508.xlsx'
    # FD_dataloader(xl_train,'20171108_New_N4ITK corrected')
    # FD_merge_dataloader()
    # assert False
    # tr_high_fdim, tr_low_fdim, tr_label_high, tr_label_low = Lac_dataloader(xl_train,'20171108_New_N4ITK corrected','../fd_result/Lac_result_v2.txt')
    # tst_high_fdim, tst_low_fdim, tst_label_high, tst_label_low = Lac_dataloader(xl_test,'Sheet1','../fd_result/Lac_result_v2.txt')
    tr_high_fdim, tr_low_fdim, tr_label_high, tr_label_low, tr_high_subj, tr_low_subj = FD_dataloader(xl_train,'20171108_New_N4ITK corrected','../fd_result/fd_result_file.txt')
    tst_high_fdim, tst_low_fdim, tst_label_high, tst_label_low, tst_high_subj, tst_low_subj = FD_dataloader(xl_test,'Sheet1','../fd_result/fd_result_file.txt')

    train_x = tr_high_fdim + tr_low_fdim
    train_y = tr_label_high + tr_label_low
    test_x =  tst_high_fdim + tst_low_fdim
    test_y =  tst_label_high + tst_label_low

    data = train_x + test_x
    label = train_y + test_y

    print(np.shape(data), np.shape(label))
    print(np.shape(data), np.shape(label))
    # assert False
    high_fd = tr_high_fdim + tst_high_fdim
    low_fd = tr_low_fdim + tst_low_fdim

    high_subj = tr_high_subj + tst_high_subj
    low_subj = tr_low_subj + tst_low_subj
    for a in high_subj:
        print(a)
    assert False

    print(np.shape(high_fd),np.shape(low_fd))
    tTestResult = stats.ttest_ind(high_fd, low_fd)
    print(tTestResult)
    print('p value : ',tTestResult[1])
    # for fd in high_fd:
    #     print(fd[:6])
    # print()
    # for fd in low_fd:
    #     print(fd[:6])

    # FD_dataloader(xl_train,'20171108_New_N4ITK corrected')

