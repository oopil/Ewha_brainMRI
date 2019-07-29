from sklearn.utils import shuffle
from pandas import get_dummies
from scipy import stats
import openpyxl
import csv
import os

import numpy as np

class MRI_chosun_data():
    def __init__(self):
        self.class_array = []
        self.nn_data = []
        self.nn_label = []
#%%
    def read_excel_data(self, excel_path):
        xl_file_name = excel_path
        xl_password = '!adai2018@#'
        xl = openpyxl.load_workbook(xl_file_name, read_only=True)
        ws = xl['Sheet1']
        self.data_excel = []
        for row in ws.rows:
            line = []
            for cell in row:
                line.append(cell.value)
            self.data_excel.append(line)
        # self.data_excel = np.array(self.data_excel)
        return self.data_excel

    def get_label_info_excel(self):
        print('Column name : ')
        print(self.data_excel[0])
        index_list = [4,5,6] # PET NEW CLINIC
        '''
        ['MRI_id', 'gender', 'age', 'education', 'amyloid PET result', 'Clinic Diagnosis', 'New Diag',
        'mtype', 'c4', ...]
        '''
        self.cnn_data = \
            [[self.data_excel[i][0],self.data_excel[i][4],self.data_excel[i][5],\
              self.data_excel[i][6],self.data_excel[i][2]]\
             for i in range(1, len(self.data_excel)) if i%3 == 0]
        print('label infomation length : {}' .format(len(self.cnn_data)))
        return self.cnn_data

    def squeeze_excel(self, excel_option):
        '''
        because there are 3 lines for each patient basically,
        we have to choose how to handle it.

        choose only one line or merge all of them
        and then remove only zero column
        '''
        print('squeeze the excel.')
        if not excel_option in self.excel_option:
            print('the excel option in not proper.')
            print(excel_option, self.excel_option)
            assert False

        option_index = self.excel_option.index(excel_option)
        print('excel option : ',excel_option, option_index)
        '''
        ['MRI_id', 'gender', 'age', 'education', 'amyloid PET result', 
        'Clinic Diagnosis', 'New Diag', 'mtype', 'c4', ...]
        '''
        for i in range(1,len(self.data_excel)):
            '''
            should test if or not demo score help the model classify correctly.
            the age factor seems to provide useful information.
            but not sure about gender and education
            '''
            # print(label_info)
            gender = [self.convert_gender_to_int(self.data_excel[i][1])]
            # demo_score = gender + self.data_excel[i][2:4] # all factors
            # demo_score = self.data_excel[i][2:4] # age + edu
            demo_score = self.data_excel[i][2:3] # only age
            # demo_score = self.data_excel[i][3:4] # only education

            if (i-1)%3 == option_index: # if choose only one row : P V T
                line = self.data_excel[i][8:]
                label_info = self.data_excel[i][4:7]
                new_line = demo_score + line
                self.nn_data.append(new_line)
                self.nn_label.append(label_info)
                # print(len(self.data_excel[i]), len(line))
                # print(new_line)

            if option_index == 3 and i%3 == 1: # if use all three rows : merge
                line = [self.data_excel[i+k][8:] for k in range(3)]
                label_info = self.data_excel[i][4:7]
                new_line = demo_score + line[0] + line[1] + line[2]
                self.nn_data.append(new_line)
                self.nn_label.append(label_info)
                # print(len(self.data_excel[i][:10]), len(line[0]), len(line[1]), len(line[2]))
                # print(new_line)
        return self.nn_data, self.nn_label

    def remove_zero_column(self):
        self.nn_data = np.array(self.nn_data)
        l1, l2 = len(self.nn_data), len(self.nn_data[0])
        delete_col_count = 0
        print('remove zero value only columns.')
        print('matrix size : ',l1, 'X', l2)
        for col in range(l2):
            is_zero_col = True
            col_index = l2 - col - 1
            for row in range(l1):
                # print(type(self.nn_data[row][4]))
                if self.nn_data[row][col_index]:
                    # print(self.nn_data[row][4])
                    # print(self.nn_data[row][col_index])
                    is_zero_col = False
                    break

            if is_zero_col:
                # print('delete column.')
                delete_col_count += 1
                self.nn_data = np.delete(self.nn_data, col_index, 1)
            # assert False
        print('removed {} columns.\t{}=>{}'.format(delete_col_count, l2, len(self.nn_data[0])))
        print('' .format() )
        return self.nn_data, self.nn_label

#%%
    def convert_gender_to_int(self, gender:str)->int:
        if gender == 'M':   return 0
        elif gender == 'F': return 1
        else:
            print('inappropriate gender is entered : ', gender)
            assert False

    def label_pet(self, label, class_array):
        for i, c in enumerate(self.class_array):
            if c in label:
                return i
                # print(c, label)
                break
        print('there is no appropriate label name. :')
        print(self.class_array, label)
        assert False

    def label_clinic(self, label, class_array):
        for i, c in enumerate(self.class_array):
            if c in label:
                '''
                print(label, i) # check the labeling state                
                '''
                return i
        if 'MCI' in label or 'AD' in label or 'CN' in label:
            return -1
        print('there is no appropriate label name. :')
        print(self.class_array, label)
        assert False

    def label_new(self, label, class_array):
        for i, c in enumerate(self.class_array):
            if c in label:
                return i
        if 'NC' in label or 'AD' in label:
            return -1
        print('there is no appropriate label name. :')
        print(self.class_array, label)
        assert False
#%%
    def set_label_func(self, label_info, class_option):
        if self.diag_type == 'PET':
            self.class_array = self.class_option_dict_pet[class_option]
            self.label_name = label_info[:,0]
            # print(self.label_name, self.class_array)
            self.label_func = self.label_pet

        elif self.diag_type == 'clinic':
            self.class_array = self.class_option_dict_clinic[class_option]
            self.label_name = label_info[:,1]
            self.label_func = self.label_clinic

        elif self.diag_type == 'new':
            self.class_array = self.class_option_dict_new[class_option]
            self.label_name = label_info[:,2]
            self.label_func = self.label_new
        else:
            print('diagnosis type is wrong. : ', self.diag_type)
            assert False
        print('class option : {} / class array : {}'.format(class_option, self.class_array))

    def remove_minus_one_label(self):
        '''
            remove the -1 line of label and data
        '''
        print('remove the -1 line of label and data.')
        self.label_list = np.array(self.label_list)
        label_length = len(self.label_list)
        for row in range(len(self.label_list)):
            row_index = label_length - row - 1
            label = self.label_list[row_index]
            # print(row, len(self.label_list), row_index)
            if label == -1:
                self.data = np.delete(self.data, row_index, 0)
                self.label_list = np.delete(self.label_list, row_index, 0)

    def define_label_nn(self, label_info, class_option):
        '''
        :param label_info: it has 3 columns : pet, new, clinic
        :param class_option:  'PET pos vs neg', 'NC vs MCI vs AD' 'NC vs mAD vs aAD vs ADD'
        :return:

        when we use the class option as NC vs AD, we need to remove MCI line.
        '''
        is_print = False
        if is_print: pass
        print('start labeling...' )
        label_info = np.array(label_info)
        self.data = self.nn_data
        self.label_list = []
        # self.class_array = self.get_class_array(class_option)
        self.set_label_func(label_info, class_option)
        for i, label in enumerate(self.label_name):
            self.label_list.append(self.label_func(label, self.class_array))

        self.remove_minus_one_label()
        if is_print:
            print(len(self.data),len(self.label_list))
            print(type(self.data),type(self.label_list))
            # print(self.nn_data[0])
            print(self.label_list)

        return self.data, self.label_list

    def define_label_cnn(self, label_info, class_option):
        '''
        :param label_info: it has 3 columns : pet, new, clinic
        :param class_option:  'PET pos vs neg', 'NC vs MCI vs AD' 'NC vs mAD vs aAD vs ADD'
        :return:

        when we use the class option as NC vs AD, we need to remove MCI line.
        '''
        is_print = True
        if is_print: pass
        print('start labeling...' )
        label_info = np.array(label_info)
        self.data = self.cnn_data
        self.label_list = []
        # self.class_array = self.get_class_array(class_option)
        self.set_label_func(label_info, class_option)
        for i, label in enumerate(self.label_name):
            self.label_list.append(self.label_func(label, self.class_array))
        self.remove_minus_one_label()
        self.label_list = np.array(self.label_list)

        if is_print:
            print('data and label length : ', len(self.data), ' = ',len(self.label_list))
            print(type(self.data),type(self.label_list))
            # print(self.nn_data[0])
            print(self.label_list)

        return self.data[:,-1], self.label_list

    def shuffle_data(self, data, label):
        print('shuffle the data and label - data length : ',len(data),' = ', len(label) )
        assert len(data)==len(label)
        random_list = [i for i in range(len(data))]
        '''
        choose between shuffle and shuffle_static
        # random_list = shuffle_static( ... )
        '''
        shuffle(random_list)
        # print(random_list)
        self.shuffle_data, self.shuffle_label = [],[]
        for index in random_list:
            self.shuffle_data.append(data[index])
            self.shuffle_label.append(label[index])

        assert len(label) == len(self.shuffle_label)
        return self.shuffle_data, self.shuffle_label

#%%
def one_hot_pd(label):
    return np.array(get_dummies(label))

def shuffle_static(arr1, arr2):
    return shuffle(arr1, arr2, random_state=0)

def valence_class(data, label, class_num):
    print('Valence the number of train and test dataset')
    length = len(data)
    label_count = [0 for i in range(class_num)]
    label_count_new = [0 for i in range(class_num)]

    for i in sorted(label):
        label_count[i] += 1

    # print('label count : ', label_count)
    min_count = min(label_count)
    print('minimun count : ',min_count)
    new_data = []
    new_label = []
    for i, k in enumerate(label):
        if label_count_new[k] < min_count:
            new_data.append(data[i])
            new_label.append(label[i])
            label_count_new[k] += 1
    # print('new label count : ', label_count_new)
    print('down sampling : {} -> {}.'.format(label_count, label_count_new))
    return np.array(new_data), np.array(new_label)

def split_data_by_fold(data, label, fold_num):
    '''
    :return: return all possible train and test set according to the fold number.
    '''
    label_set = list(set(label))
    print('split the data into train and test by fold number. fold number : {} label set :{}' \
          .format(fold_num, label_set))
    separate_data = [[] for _ in range(len(label_set))]
    separate_label = [[] for _ in range(len(label_set))]
    # separate the data into different label list
    for i, l in enumerate(label):
        # print(i,l,label_count)
        separate_data[l].append(data[i])
        separate_label[l].append(label[i])
    # print(separate_label)
    label_count = [len(i) for i in separate_label]
    test_count = [count // fold_num for count in label_count]
    print(label_count, test_count)
    smaller_data_num = min(label_count)
    test_num = smaller_data_num // fold_num

    whole_set = []
    for fold_index in range(fold_num):
        tr_x, tr_y, tst_x, tst_y = [], [], [], []
        for i, one_label in enumerate(separate_data):
            if fold_index == fold_num - 1:
                tr_x = tr_x + one_label[:test_count[i] * fold_index]
                tr_y = tr_y + separate_label[i][:test_count[i] * fold_index]
                tst_x = tst_x + one_label[test_count[i] * fold_index:]
                tst_y = tst_y + separate_label[i][test_count[i] * fold_index:]
                pass
            else:
                tr_x = tr_x + one_label[:test_count[i] * fold_index] + one_label[test_count[i] * (
                            fold_index + 1):]
                tr_y = tr_y + separate_label[i][:test_count[i] * fold_index] + separate_label[i][
                                                                                             test_count[i] * (
                                                                                                         fold_index + 1):]
                tst_x = tst_x + one_label[test_count[i] * fold_index:test_count[i] * (fold_index + 1)]
                tst_y = tst_y + separate_label[i][
                                          test_count[i] * fold_index:test_count[i] * (fold_index + 1)]
        print(len(tr_y) + len(tst_y), len(tr_y), len(tst_y))
        tr_x = np.array(tr_x)
        tst_x = np.array(tst_x)
        tr_y = np.array(tr_y)
        tst_y = np.array(tst_y)
        whole_set.append([tr_x, tr_y, tst_x, tst_y])
    return whole_set

def normalize_col(X_, axis=0):
    """
    if minimun is negative float, then we should divide with min + max
    :param X_:
    :param axis:
    :return:
    """
    print("normalize step")
    # print(X_[:2])
    # print(np.amax(X_[:2],axis=axis))
    assert len(np.amax(X_,axis=axis)) == len(X_[0])
    assert np.all(np.amax(X_,axis=axis) != 0)
    # print(np.amin(X_,axis=axis)[41])
    # print(np.amax(X_, axis=axis)[41])
    # print(((X_-np.amin(X_,axis=axis))/np.amax(X_, axis=axis))[0,41])
    X_ = X_-np.amin(X_,axis=axis)
    return (X_-np.amin(X_,axis=axis))/np.amax(X_, axis=axis)

def column(matrix, i, num):
    col = [row[i:i+num] for row in matrix]
    for row in matrix:
        if row[i:i+num] == []:
            print('Here is the criminal')
            print(row)
            assert False
    return col

def NN_dataloader(diag_type, class_option, \
                  excel_path, excel_option, test_num, fold_num, is_split_by_num):
    '''
    1. read excel data (O)
    2. squeeze 3 lines into 1 lines according to the options P V T merge (O)
    5. normalization -> should do this at first. separately according to the column.
    no. we don't have to do this at first. because column is preserved.
    3. remove zero value only column (O)
    3. make label list (O)
    4. shuffle (O)
    6. split train and test dataset (O)
    :return: train and test data and lable
    '''
    # "clinic" or "new" or "PET"
    # 'PET pos vs neg', 'NC vs MCI vs AD' 'NC vs mAD vs aAD vs ADD'
    # diag_type = "PET"
    # class_option = 'PET pos vs neg'
    # diag_type = "new"
    # class_option = 'NC vs mAD vs aAD vs ADD'
    # diag_type = "clinic"
    # class_option = 'NC vs AD' #'NC vs MCI vs AD'
    # base_folder_path = '/home/sp/Datasets/MRI_chosun/ADAI_MRI_Result_V1_0'
    # base_folder_path = '/home/sp/Datasets/MRI_chosun/test_sample_2'
    # excel_path = '/home/sp/Datasets/MRI_chosun/ADAI_MRI_test.xlsx'
    # excel_option = 'merge' # P V T merge

    loader = MRI_chosun_data()
    loader.set_diagnosis_type(diag_type)
    loader.read_excel_data(excel_path)
    loader.squeeze_excel(excel_option=excel_option)
    data, label_info = loader.remove_zero_column()

    data, label = loader.define_label_nn(label_info, class_option)
    # normalize each column separately.
    data = normalize_col(data)
    # data = normalize_separate_col(data)
    # print(data.shape)
    # data = normalize(data)
    '''
    when split the data by fold number, should we split the data earlier than shuffle??
    '''
    # is_split_by_num = False
    if is_split_by_num:
        shuffle_data, shuffle_label = loader.shuffle_data(data, label)
        # test_num = 20
        '''
            return only one train and test set
        '''
        return loader.split_data_by_num(shuffle_data, shuffle_label, test_num)
    else:
        shuffle_data, shuffle_label = loader.shuffle_data(data, label)
        # fold_num = 5
        # fold_index = 0
        '''
            return all train and test sets devided by fold. 
        '''
        return loader.split_data_by_fold(shuffle_data, shuffle_label, fold_num)

from imblearn.over_sampling import *
from imblearn.combine import *

def over_sampling(X_imb, Y_imb, sampling_option):
    print('starts over sampling ...', sampling_option)
    is_reshape = False
    shape = X_imb.shape
    if np.ndim(X_imb) == 1 and sampling_option != 'SIMPLE':
        is_reshape = True
        X_imb = X_imb.reshape(-1,1)
        print(X_imb.shape)

    if sampling_option == 'ADASYN':
        X_samp, Y_samp = ADASYN().fit_sample(X_imb, Y_imb)
    elif sampling_option == 'SMOTE':
        X_samp, Y_samp = SMOTE(random_state=4).fit_sample(X_imb, Y_imb)
    elif sampling_option == 'SMOTEENN':
        X_samp, Y_samp = SMOTEENN(random_state=0).fit_sample(X_imb, Y_imb)
    elif sampling_option == 'SMOTETomek':
        X_samp, Y_samp = SMOTETomek(random_state=4).fit_sample(X_imb, Y_imb)
    elif sampling_option == 'RANDOM':
        X_samp, Y_samp = RandomOverSampler(random_state=0).fit_sample(X_imb, Y_imb)
    elif sampling_option == 'BolderlineSMOTE':
        X_samp, Y_samp = BorderlineSMOTE().fit_sample(X_imb, Y_imb)

    elif sampling_option == 'SIMPLE':
        label_set = list(set(Y_imb))
        separate_data = [[] for _ in range(len(label_set))]
        separate_label = [[] for _ in range(len(label_set))]
        separate_length = []
        # separate the data into different label list
        for i, l in enumerate(Y_imb):
            # print(i,l,label_count)
            separate_data[l].append(X_imb[i])
            separate_label[l].append(Y_imb[i])

        print('sampling option is just scaling the number.')
        for l in separate_label:
            print(len(l), end=' / ')
            separate_length.append(len(l))
        print()
        mini = np.min(separate_length)
        maxi = np.max(separate_length)
        max_index = separate_length.index(maxi)
        scale = np.round(maxi/mini)
        print(mini, maxi,' => ' , scale)
        print('max index : ',separate_length.index(maxi))

        X_samp = separate_data[max_index]
        Y_samp = separate_label[max_index]
        print(np.shape(X_samp), np.shape(Y_samp))
        for c in range(len(separate_label)):
            if c != max_index:
                scale = int(np.round(maxi/len(separate_label[c])))
                X_samp = X_samp + separate_data[c]*scale
                Y_samp = Y_samp + separate_label[c]*scale
        X_samp =np.array(X_samp)
        Y_samp = np.array(Y_samp)
    elif sampling_option == 'None':
        X_samp, Y_samp = X_imb, Y_imb
    else:
        print('sampling option is not proper.', sampling_option)
        assert False

    if is_reshape:
        X_samp = np.squeeze(X_samp)
        print(X_samp.shape)
        # assert False

    imbalance_num = len(Y_imb)
    balance_num = len(Y_samp)
    print('over sampling from {:5} -> {:5}.'.format(imbalance_num, balance_num))
    return X_samp, Y_samp

def read_csv(file_path)->np.array:
    f = open(file_path, 'r', encoding='utf-8')
    rdr = csv.reader(f)
    contents = []
    for line in rdr:
        contents.append(line)
        # print(len(line),line)
    f.close()
    return np.array(contents)

def check_nan(l:list)->bool:
    is_ = False
    for i, e in enumerate(l):
        if len(str(e)) == 0:
            print(i,e,end=' ')
            is_ = True
    if is_:
        print()
    return is_

def EWHA_excel_datareader():
    base_folder_path = '/home/soopil/Desktop/Dataset/brain_ewha'# desktop setting
    train_path = os.path.join(base_folder_path, 'Train_Meningioma_20180508.xlsx')
    test_path = os.path.join(base_folder_path, 'Test_Meningioma_20180508.xlsx')

    xl = openpyxl.load_workbook(train_path, read_only=True)
    ws = xl['20171108_New_N4ITK corrected']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)
    train_x = data_excel[1:, 4:102]
    train_y = data_excel[1:, 2]
    # for i,line in enumerate(train_x):
    #     print(i, np.where(line == 0))
    train_x = np.delete(train_x, 109, 0)
    train_y = np.delete(train_y, 109, 0)
    # for line in train_x:
    #     print(np.where(line>1))
    # assert False
    xl = openpyxl.load_workbook(test_path, read_only=True)
    ws = xl['Sheet1']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)

    test_x = data_excel[1:, 4:102]
    test_y = data_excel[1:, 2]
    # for i,line in enumerate(test_x):
    #     print(i, np.where(line == 0))
    test_x = np.delete(test_x, 59, 0)
    test_y = np.delete(test_y, 59, 0)

    is_norm = True
    if is_norm:
        train_x = normalize_col(train_x, axis=0)
        test_x = normalize_col(test_x, axis=0)

    print(train_x.shape,train_y.shape, test_x.shape,test_y.shape)
    # assert False
    return train_x, train_y, test_x, test_y

def EWHA_external_datareader():
    # --------------------- excel file read --------------------- #
    base_folder_path = '/home/soopil/Desktop/Dataset/brain_ewha/external_validation'# desktop setting
    external_path = os.path.join(base_folder_path,'External validation set_Ewha_Meningioma.xlsx')
    xl = openpyxl.load_workbook(external_path, read_only=True)
    ws = xl['Sheet1']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)

    data_excel = np.array(data_excel)
    subj_list_excel= data_excel[1:,0]
    # print(subj_list_excel)
    # assert False
    print(data_excel[0])
    # print(len(data_excel[1:,0]))
    label_list = np.squeeze(data_excel[1:, 5:6])

    # print(np.shape(label_list))
    # print(label_list)
    # print(len(np.where(label_list == 1)[0]),len(np.where(label_list == 0)[0]))
    # assert False
    # --------------------- csv file read --------------------- #
    subj_list_csv = []
    csv_dir_path = os.path.join(base_folder_path, 'CSV')
    csv_file_list = os.listdir(csv_dir_path)
    for i, e in enumerate(sorted(csv_file_list)):
        e_split = e.split('_')
        subj_name = e_split[0]
        modality = e_split[-1]#.split('.')[0]
        subj_list_csv.append(int(subj_name))
        print(subj_name, modality)

        csv_path = os.path.join(csv_dir_path, e)
        contents = read_csv(csv_path)

        # from 38 line Voxel volume
        feature_name = contents[:,0]
        # print(np.where(feature_name == 'original'))
        i_start = np.where(feature_name == 'original')[0][0]
        i_end = np.where(feature_name == 'original')[0][-1]
        length = i_end-i_start
        # print(i_start, i_end, length)
        if subj_name == '11311865' and modality == 'T1C.csv':
            features = contents[i_start:i_start+length, 4].astype(np.float32)
        else:
            features =  contents[i_start:i_start+length:, 3].astype(np.float32)

        # print(len(features))
        assert length == 106

        if check_nan(features):
            print(e)

            # assert False
        # print(features)
        # print(contents)
        # print(np.shape(contents), np.shape(features))
        #
        # assert False
    print('csv total count : ',len(csv_file_list))

    # --------------------- MRI image file read --------------------- #
    print('<< check the MRI image files >>')
    MRI_dir_path_1 = os.path.join(base_folder_path,'External_Meningioma_T1C_1')
    MRI_dir_path_2 = os.path.join(base_folder_path,'External_Meningioma_T1C_2')
    MRI_label_path = os.path.join(base_folder_path,'label')
    MRI_dir_list_1 = os.listdir(MRI_dir_path_1)
    MRI_dir_list_2 = os.listdir(MRI_dir_path_2)
    MRI_label_list = os.listdir(MRI_label_path)

    print(len(MRI_dir_list_1))
    print(len(MRI_dir_list_2))
    print(len(MRI_label_list))

    total_list = MRI_dir_list_1 + MRI_dir_list_2
    print(len(set(total_list)))
    # print(MRI_dir_list_2)

    print(set(subj_list_csv))
    print(set(subj_list_excel))
    print(set(subj_list_csv) == set(subj_list_excel))

    tmp = []
    for e in total_list:
        tmp.append(int(e.split('_')[0]))

    for a,b in zip(sorted(set(subj_list_csv)), sorted(tmp)):
        print(a,b)
        assert a == b

def check_features():
    '''
    print all the features in CSV files compared to the original excel data.
    :return:
    '''
    # --------------------- ER meningloma file read --------------------- #
    base_folder_path = '/home/soopil/Desktop/Dataset/brain_ewha'  # desktop setting
    train_path = os.path.join(base_folder_path, 'Train_Meningioma_20180508.xlsx')
    test_path = os.path.join(base_folder_path, 'Test_Meningioma_20180508.xlsx')
    xl = openpyxl.load_workbook(train_path, read_only=True)
    ws = xl['20171108_New_N4ITK corrected']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)
    # train_x = data_excel[1:, 4:102]
    feature_name_list = data_excel[0, 4:102]
    feature_T1 = []
    for e in feature_name_list:
        if 'T1C' in e:
            feature_T1.append(e)
            # print(e)
    # assert False
    # print(np.where('T1C' in feature_name_list))

    # --------------------- external file read --------------------- #
    base_folder_path = '/home/soopil/Desktop/Dataset/brain_ewha/external_validation'  # desktop setting
    external_path = os.path.join(base_folder_path, 'External validation set_Ewha_Meningioma.xlsx')
    xl = openpyxl.load_workbook(external_path, read_only=True)
    ws = xl['Sheet1']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)

    data_excel = np.array(data_excel)
    # print(data_excel[0,10:])
    feature_name_list_external = data_excel[0,10:]
    feature_T1_external = []
    for e in feature_name_list_external:
        if 'T1C' in e:
            feature_T1_external.append(e)
            # print(e)
    print(feature_T1)
    print(feature_T1_external)

    # for a,b in zip(feature_T1, feature_T1_external):
    #     print('{:30} | {:30}'.format(a,b) )

    l1, l2 = len(feature_T1), len(feature_T1_external)
    length = np.max([l1,l2])
    print(length)

    print('<{:30}> <{:30}>'.format('existing T1 features', 'features from external dataset'))
    print()
    for i in range(length):
        if i >= l1:
            a = ' - '
        else:
            a = feature_T1[i]
        if i >= l2:
            b = ' - '
        else:
            b = feature_T1_external[i]

        print('{:30} | {:30}'.format(a,b) )

def EWHA_CSV_reader():
    # --------------------- excel file read --------------------- #
    base_folder_path = "/home/soopil/Desktop/Dataset/EWHA_brain_tumor/0_Fwd_ MENINGIOMA 추가 자료 1_190711"
    external_path = os.path.join(base_folder_path,'Training set_Sinchon_Meningioma.xlsx')
    xl = openpyxl.load_workbook(external_path, read_only=True)
    ws = xl['20171108_New_N4ITK corrected']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)

    data_excel = np.array(data_excel)
    subj_list_excel= data_excel[1:,0]
    label_list = np.squeeze(data_excel[1:, 5:5+1])
    sex_list = np.squeeze(data_excel[1:, 107:107+1])
    print(len(subj_list_excel),subj_list_excel)
    print(data_excel[0])
    print(label_list)
    print(sex_list)
    print('instance number of each class : ',len(np.where(label_list == 1)[0]),len(np.where(label_list == 0)[0]))

    # --------------------- Fractal dimention result file read --------------------- #
    result_dpath = '/home/soopil/Desktop/github/Ewha_brainMRI/fd_result/'
    # result_fname = 'SINCHON_FD_result_20190718.txt'
    # result_fname = 'SINCHON_FD_result_20190719_rescale.txt'
    result_fname = 'SINCHON_FD_result_20190723.txt' # original dataset
    result_fpath = os.path.join(result_dpath, result_fname)

    fd = open(result_fpath)
    lines = fd.readlines()
    print(lines)

    def FD(box_count):
        box_count = np.array(box_count).astype(np.float32)
        length = len(box_count)
        r = np.array([2 ** i for i in range(0, length)]).astype(np.float32)
        n_grad = np.gradient(np.log(box_count))
        r_grad = np.gradient(np.log(r))
        return (np.divide(n_grad, r_grad) * (-1))

    subj_list_fd, data_fd = [],[]

    for line in lines:

        if line == 'box counting fractal dimension.\n':
            continue
        split = line.split('/')
        subj = split[0]
        bx_count = split[2].split(',')
        lac = split[3].split(',')
        lac = list(np.array(lac).astype(np.float32))
        fdim = list(FD(bx_count))
        print(fdim+lac)
        assert len(fdim) == 10 and len(lac) == 9
        subj_list_fd.append(subj)
        data_fd.append(fdim+lac)
        # print(subj, bx_count, lac)
        # print(len(bx_count), len(lac))
        # print(list(FD(np.flip(lac)))) # ???? let me try
        # print(list(fdim))
        print()
    fd.close()
    print(len(subj_list_fd))
    # assert False
    # need certain normalization or other processes
    # ===================== normalize FD features ===================== #
    is_norm = False
    if is_norm:
        # normalize_col()
        data_fd = np.array(data_fd)
        maxi = np.amax(data_fd[:,:10])
        mini = np.amin(data_fd[:,:10])
        data_fd[:,:10] = (data_fd[:,:10] - mini) / maxi
        print(maxi, mini)
        # for e in data_fd:
        #     print(e)
        # assert False

    # ===================== csv file read ===================== #
    data_f, label_f  = [],[]
    subj_list_csv = []
    subj_list_no_csv = []
    csv_dir_path = os.path.join(base_folder_path, 'CSV')
    csv_file_list = os.listdir(csv_dir_path)
    print('total csv file number : ',len(csv_file_list))
    print(csv_file_list)

    for i, subj in enumerate(subj_list_excel):
        label, sex = label_list[i], sex_list[i]

        if sex == 'M':
            sex = [0.]
        else:
            sex = [1.]

        # print(subj, label, sex)
        T1C = str(subj)+'_T1C.csv'
        T2 = str(subj)+'_T2.csv'

        if (str(subj) not in subj_list_csv) and (str(subj) in subj_list_fd) or (subj in subj_list_fd):
            # sex features label
            is_FD_set = True #False
            # features = sex + list(features_T1C) + list(features_T2)
            features = []
            if is_FD_set:
                index_fd = subj_list_fd.index(str(subj))
                fd = list(data_fd[index_fd][:10]) #10 is fd and 9 is LAC
                features = features + fd

            # print(features)
            subj_list_csv.append(subj)
            data_f.append(features)
            label_f.append(label)
        else:
            subj_list_no_csv.append(subj)

    print('subjects with CSV files : ', len(subj_list_csv), subj_list_csv)
    print('subjects with no CSV files : ', len(subj_list_no_csv), subj_list_no_csv)
    return subj_list_csv, data_f, label_f

def EWHA_CSV_reader_save():
    # --------------------- excel file read --------------------- #
    base_folder_path = "/home/soopil/Desktop/Dataset/EWHA_brain_tumor/0_Fwd_ MENINGIOMA 추가 자료 1_190711"
    external_path = os.path.join(base_folder_path,'Training set_Sinchon_Meningioma.xlsx')
    xl = openpyxl.load_workbook(external_path, read_only=True)
    ws = xl['20171108_New_N4ITK corrected']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)

    data_excel = np.array(data_excel)
    subj_list_excel= data_excel[1:,0]
    label_list = np.squeeze(data_excel[1:, 5:5+1])
    sex_list = np.squeeze(data_excel[1:, 107:107+1])
    print(len(subj_list_excel),subj_list_excel)
    print(data_excel[0])
    print(label_list)
    print(sex_list)
    print('instance number of each class : ',len(np.where(label_list == 1)[0]),len(np.where(label_list == 0)[0]))

    # --------------------- Fractal dimention result file read --------------------- #
    result_dpath = '/home/soopil/Desktop/github/Ewha_brainMRI/fd_result/'
    # result_fname = 'SINCHON_FD_result_20190718.txt'
    # result_fname = 'SINCHON_FD_result_20190719_rescale.txt'
    result_fname = 'SINCHON_FD_result_20190723.txt' # original dataset
    result_fpath = os.path.join(result_dpath, result_fname)

    fd = open(result_fpath)
    lines = fd.readlines()
    print(lines)

    def FD(box_count):
        box_count = np.array(box_count).astype(np.float32)
        length = len(box_count)
        r = np.array([2 ** i for i in range(0, length)]).astype(np.float32)
        n_grad = np.gradient(np.log(box_count))
        r_grad = np.gradient(np.log(r))
        return (np.divide(n_grad, r_grad) * (-1))

    subj_list_fd, data_fd = [],[]

    for line in lines:

        if line == 'box counting fractal dimension.\n':
            continue
        split = line.split('/')
        subj = split[0]
        bx_count = split[2].split(',')
        lac = split[3].split(',')
        lac = list(np.array(lac).astype(np.float32))
        fdim = list(FD(bx_count))
        print(fdim+lac)
        assert len(fdim) == 10 and len(lac) == 9
        subj_list_fd.append(subj)
        data_fd.append(fdim+lac)
        # print(subj, bx_count, lac)
        # print(len(bx_count), len(lac))
        # print(list(FD(np.flip(lac)))) # ???? let me try
        # print(list(fdim))
        print()
    fd.close()

    # need certain normalization or other processes
    # ===================== normalize FD features ===================== #
    is_norm = False
    if is_norm:
        # normalize_col()
        data_fd = np.array(data_fd)
        maxi = np.amax(data_fd[:,:10])
        mini = np.amin(data_fd[:,:10])
        data_fd[:,:10] = (data_fd[:,:10] - mini) / maxi
        print(maxi, mini)
        # for e in data_fd:
        #     print(e)
        # assert False

    # ===================== csv file read ===================== #
    data_f, label_f  = [],[]
    subj_list_csv = []
    subj_list_no_csv = []
    csv_dir_path = os.path.join(base_folder_path, 'CSV')
    csv_file_list = os.listdir(csv_dir_path)
    print('total csv file number : ',len(csv_file_list))
    print(csv_file_list)

    for i, subj in enumerate(subj_list_excel):
        label, sex = label_list[i], sex_list[i]

        if sex == 'M':
            sex = [0.]
        else:
            sex = [1.]

        # print(subj, label, sex)
        T1C = str(subj)+'_T1C.csv'
        T2 = str(subj)+'_T2.csv'

        if (str(subj) in subj_list_fd) and (T1C in csv_file_list) and (T2 in csv_file_list):
            '''
            only use files with mask, T1C csv, T2C csv
            need to normalize features from T1C and T2 here
            '''
            # T1C csv file read
            csv_path = os.path.join(csv_dir_path, T1C)
            contents = read_csv(csv_path)
            # from 38 line Voxel volume
            feature_name = contents[:, 0]
            i_start = np.where(feature_name == 'original')[0][0]
            i_end = np.where(feature_name == 'original')[0][-1]
            length = i_end - i_start
            assert length == 106
            features_T1C = contents[i_start:i_start + length:, 3].astype(np.float32)

            if check_nan(features_T1C):
                print('detect nan number in T1C features : ',subj)

            # T2 csv file read
            csv_path = os.path.join(csv_dir_path, T2)
            contents = read_csv(csv_path)

            # format exception => skip first
            if subj in [5596598, 8453071]:
                subj_list_no_csv.append(subj)
                continue
                f = open(csv_path, 'r', encoding='utf-8')
                rdr = csv.reader(f)
                assert False

            # from 38 line Voxel volume
            feature_name = contents[:, 0]
            i_start = np.where(feature_name == 'original')[0][0]
            i_end = np.where(feature_name == 'original')[0][-1]
            length = i_end - i_start
            assert length == 106
            features_T2 = contents[i_start:i_start + length:, 3].astype(np.float32)

            if check_nan(features_T2):
                print('detect nan number in T2 features : ', subj)

            # sex features label
            is_FD_set = True #False
            # features = sex + list(features_T1C) + list(features_T2)
            features = []
            if is_FD_set:
                index_fd = subj_list_fd.index(str(subj))
                fd = list(data_fd[index_fd][:10]) #10 is fd and 9 is LAC
                features = features + fd

            # print(features)
            subj_list_csv.append(subj)
            data_f.append(features)
            label_f.append(label)
        else:
            subj_list_no_csv.append(subj)

    print('subjects with CSV files : ', len(subj_list_csv), subj_list_csv)
    print('subjects with no CSV files : ', len(subj_list_no_csv), subj_list_no_csv)
    return subj_list_csv, data_f, label_f

def read_xl(xl_path, sheet_name):
    xl = openpyxl.load_workbook(xl_path, read_only=True)
    ws = xl[sheet_name]
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)
    print(data_excel.shape)
    subj_list = data_excel[1:, 0]
    label_list = data_excel[1:, 2]
    print(label_list)
    assert len(subj_list) == len(label_list)
    return subj_list, label_list

def SINCHON_FD_reader():
    # --------------------- excel file read --------------------- #
    # xl_test = '/home/soopil/Desktop/Dataset/brain_ewha_early/Test_Meningioma_20180508.xlsx'
    # xl_train = '/home/soopil/Desktop/Dataset/brain_ewha_early/Train_Meningioma_20180508.xlsx'
    xl_path = '/home/soopil/Desktop/Dataset/brain_ewha_early/20180508_Entire_Train and Test_Meningioma.xlsx'
    # base_folder_path = "/home/soopil/Desktop/Dataset/EWHA_brain_tumor/0_Fwd_ MENINGIOMA 추가 자료 1_190711"
    # xl_path = os.path.join(base_folder_path,'Training set_Sinchon_Meningioma.xlsx')

    subj_list_excel, label_list = read_xl(xl_path, '20171108_New_N4ITK corrected')
    print(len(subj_list_excel))
    assert len(subj_list_excel) == len(label_list)
    # assert False
    # xl = openpyxl.load_workbook(external_path, read_only=True)
    # ws = xl['20171108_New_N4ITK corrected']
    # data_excel = []
    # for row in ws.rows:
    #     line = []
    #     for cell in row:
    #         line.append(cell.value)
    #     data_excel.append(line)
    #
    # data_excel = np.array(data_excel)
    # subj_list_excel= data_excel[1:,0]
    # label_list = np.squeeze(data_excel[1:, 5:5+1])
    # sex_list = np.squeeze(data_excel[1:, 107:107+1])
    # print(len(subj_list_excel),subj_list_excel)
    # print(data_excel[0])
    # print(label_list)
    # print(sex_list)
    # print('instance number of each class : ',len(np.where(label_list == 1)[0]),len(np.where(label_list == 0)[0]))

    # --------------------- Fractal dimention result file read --------------------- #
    result_dpath = '/home/soopil/Desktop/github/Ewha_brainMRI/fd_result/'
    # result_fname = 'SINCHON_FD_result_20190718.txt'
    # result_fname = 'SINCHON_FD_result_20190719_rescale.txt'
    result_fname = 'SINCHON_FD_result_20190723.txt' # original dataset
    result_fpath = os.path.join(result_dpath, result_fname)

    fd = open(result_fpath)
    lines = fd.readlines()
    print(lines)

    # ===================== read FD features ===================== #
    def FD(box_count):
        box_count = np.array(box_count).astype(np.float32)
        length = len(box_count)
        r = np.array([2 ** i for i in range(0, length)]).astype(np.float32)
        n_grad = np.gradient(np.log(box_count))
        r_grad = np.gradient(np.log(r))
        return (np.divide(n_grad, r_grad) * (-1))

    subj_list_fd, data_fd = [],[]

    for line in lines:

        if line == 'box counting fractal dimension.\n':
            continue
        split = line.split('/')
        subj = split[0]
        bx_count = split[2].split(',')
        lac = split[3].split(',')
        lac = list(np.array(lac).astype(np.float32))
        fdim = list(FD(bx_count))
        # lac = list(FD( np.flip(lac)))
        lac = list(np.log(lac))
        # print(fdim+lac)
        # print(lac)
        assert len(fdim) == 10 and len(lac) == 9
        subj_list_fd.append(subj)
        data_fd.append(fdim+lac)
        # print(subj, bx_count, lac)
        # print(len(bx_count), len(lac))
        # print(list(FD(np.flip(lac)))) # ???? let me try
        # print(list(fdim))
    fd.close()
    # print(len(subj_list_fd),subj_list_fd)
    # assert False

    # need certain normalization or other processes
    # ===================== normalize FD features ===================== #
    is_norm = False
    if is_norm:
        # normalize_col()
        data_fd = np.array(data_fd)
        maxi = np.amax(data_fd[:,:10])
        mini = np.amin(data_fd[:,:10])
        data_fd[:,:10] = (data_fd[:,:10] - mini) / maxi
        print(maxi, mini)
        # for e in data_fd:
        #     print(e)
        # assert False

        data_fd = np.array(data_fd)
        maxi = np.amax(data_fd[:, 10:])
        mini = np.amin(data_fd[:, 10:])
        data_fd[:, 10:] = (data_fd[:, 10:] - mini) / maxi

    # ===================== merge FD with Excel ===================== #
    data_f, label_f  = [],[]
    subj_list_f, subj_list_no_f, subj_dup  = [], [], []

    # csv_dir_path = os.path.join(base_folder_path, 'CSV')
    # csv_file_list = os.listdir(csv_dir_path)
    # print('total csv file number : ',len(csv_file_list))
    # print(csv_file_list)

    for i, subj in enumerate(sorted(subj_list_excel)):
        # label, sex = label_list[i], sex_list[i]
        label_index = np.where(subj_list_excel == subj)
        label = label_list[label_index][0]
        # if sex == 'M':
        #     sex = [0.]
        # else:
        #     sex = [1.]

        if (str(subj) in subj_list_fd) :
            # print(subj)
            if subj in subj_list_f:
                subj_dup.append(subj)
                continue

            is_FD_set = True #False
            features = []

            if is_FD_set:
                index_fd = subj_list_fd.index(str(subj))
                fd = list(data_fd[index_fd][:]) #10 is fd and 9 is LAC
                features = features + fd

            # print(subj, label, features)
            subj_list_f.append(subj)
            data_f.append(features)
            label_f.append(label)

        else:
            # print(subj)
            subj_list_no_f.append(subj)

    set_a = set(subj_list_fd) - set(subj_list_excel)
    set_b = set(subj_list_excel) - set(subj_list_fd)
    print(len(set_a), set_a)
    print(len(set_b), set_b)
    print('fd subjects : ',len(subj_list_fd), len(set(subj_list_fd)), subj_list_fd)
    print('subjects with fd files : ', len(subj_list_f), subj_list_f)
    print('subjects with no files : ', len(subj_list_no_f), subj_list_no_f)
    print('duplicated subjects : ', len(subj_dup), subj_dup)
    return subj_list_f, data_f, label_f

if __name__ == '__main__':
    # EWHA_excel_datareader()
    # EWHA_external_datareader()
    # check_features()
    # subj_list, data, label = EWHA_CSV_reader()
    subj_list, data, label = SINCHON_FD_reader()

    # for e in subj_list:
    #     print(e)
    # assert False

    low, high = [], []
    subj_low, subj_high = [], []
    for i in range(len(label)):
        if label[i] == 0:
            subj_low.append(subj_list[i])
            low.append(data[i])
        elif label[i] == 1:
            subj_high.append(subj_list[i])
            high.append(data[i])

    tTestResult = stats.ttest_ind(low, high)
    print(tTestResult[0])
    print(list(tTestResult[1]))
    print(high)

    # subj_list, data, label = SINCHON_FD_reader()
    # print(label)
    print('final class count : ',len(label), np.sum(label), len(label) - np.sum(label))
    # whole_set = split_data_by_fold(data, label, 5)
