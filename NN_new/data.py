from sklearn.utils import shuffle
from pandas import get_dummies
from scipy import stats
import numpy as np
import openpyxl
import csv
import sys
import os

sys.path.append('..')
from excel_data_reader import *

def read_csv(file_path)->np.array:
    f = open(file_path, 'r', encoding='utf-8')
    rdr = csv.reader(f)
    contents = []
    for line in rdr:
        contents.append(line)
        # print(len(line),line)
    f.close()
    return np.array(contents)

def check_nan(l:list)->bool:
    is_ = False
    for i, e in enumerate(l):
        if len(str(e)) == 0:
            print(i,e,end=' ')
            is_ = True
    if is_:
        print()
    return is_

def EWHA_excel_datareader():
    base_folder_path = '/home/soopil/Desktop/Dataset/brain_ewha'# desktop setting
    train_path = os.path.join(base_folder_path, 'Train_Meningioma_20180508.xlsx')
    test_path = os.path.join(base_folder_path, 'Test_Meningioma_20180508.xlsx')

    xl = openpyxl.load_workbook(train_path, read_only=True)
    ws = xl['20171108_New_N4ITK corrected']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)
    train_x = data_excel[1:, 4:102]
    train_y = data_excel[1:, 2]
    # for i,line in enumerate(train_x):
    #     print(i, np.where(line == 0))
    train_x = np.delete(train_x, 109, 0)
    train_y = np.delete(train_y, 109, 0)
    # for line in train_x:
    #     print(np.where(line>1))
    # assert False
    xl = openpyxl.load_workbook(test_path, read_only=True)
    ws = xl['Sheet1']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)

    test_x = data_excel[1:, 4:102]
    test_y = data_excel[1:, 2]
    # for i,line in enumerate(test_x):
    #     print(i, np.where(line == 0))
    test_x = np.delete(test_x, 59, 0)
    test_y = np.delete(test_y, 59, 0)

    is_norm = True
    if is_norm:
        train_x = normalize_col(train_x, axis=0)
        test_x = normalize_col(test_x, axis=0)

    print(train_x.shape,train_y.shape, test_x.shape,test_y.shape)
    # assert False
    return train_x, train_y, test_x, test_y

def EWHA_external_datareader():
    # --------------------- excel file read --------------------- #
    base_folder_path = '/home/soopil/Desktop/Dataset/brain_ewha/external_validation'# desktop setting
    external_path = os.path.join(base_folder_path,'External validation set_Ewha_Meningioma.xlsx')
    xl = openpyxl.load_workbook(external_path, read_only=True)
    ws = xl['Sheet1']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)

    data_excel = np.array(data_excel)
    subj_list_excel= data_excel[1:,0]
    # print(subj_list_excel)
    # assert False
    print(data_excel[0])
    # print(len(data_excel[1:,0]))
    label_list = np.squeeze(data_excel[1:, 5:6])

    # print(np.shape(label_list))
    # print(label_list)
    # print(len(np.where(label_list == 1)[0]),len(np.where(label_list == 0)[0]))
    # assert False
    # --------------------- csv file read --------------------- #
    subj_list_csv = []
    csv_dir_path = os.path.join(base_folder_path, 'CSV')
    csv_file_list = os.listdir(csv_dir_path)
    for i, e in enumerate(sorted(csv_file_list)):
        e_split = e.split('_')
        subj_name = e_split[0]
        modality = e_split[-1]#.split('.')[0]
        subj_list_csv.append(int(subj_name))
        print(subj_name, modality)

        csv_path = os.path.join(csv_dir_path, e)
        contents = read_csv(csv_path)

        # from 38 line Voxel volume
        feature_name = contents[:,0]
        # print(np.where(feature_name == 'original'))
        i_start = np.where(feature_name == 'original')[0][0]
        i_end = np.where(feature_name == 'original')[0][-1]
        length = i_end-i_start
        # print(i_start, i_end, length)
        if subj_name == '11311865' and modality == 'T1C.csv':
            features = contents[i_start:i_start+length, 4].astype(np.float32)
        else:
            features =  contents[i_start:i_start+length:, 3].astype(np.float32)

        # print(len(features))
        assert length == 106

        if check_nan(features):
            print(e)

            # assert False
        # print(features)
        # print(contents)
        # print(np.shape(contents), np.shape(features))
        #
        # assert False
    print('csv total count : ',len(csv_file_list))

    # --------------------- MRI image file read --------------------- #
    print('<< check the MRI image files >>')
    MRI_dir_path_1 = os.path.join(base_folder_path,'External_Meningioma_T1C_1')
    MRI_dir_path_2 = os.path.join(base_folder_path,'External_Meningioma_T1C_2')
    MRI_label_path = os.path.join(base_folder_path,'label')
    MRI_dir_list_1 = os.listdir(MRI_dir_path_1)
    MRI_dir_list_2 = os.listdir(MRI_dir_path_2)
    MRI_label_list = os.listdir(MRI_label_path)

    print(len(MRI_dir_list_1))
    print(len(MRI_dir_list_2))
    print(len(MRI_label_list))

    total_list = MRI_dir_list_1 + MRI_dir_list_2
    print(len(set(total_list)))
    # print(MRI_dir_list_2)

    print(set(subj_list_csv))
    print(set(subj_list_excel))
    print(set(subj_list_csv) == set(subj_list_excel))

    tmp = []
    for e in total_list:
        tmp.append(int(e.split('_')[0]))

    for a,b in zip(sorted(set(subj_list_csv)), sorted(tmp)):
        print(a,b)
        assert a == b

def check_features():
    '''
    print all the features in CSV files compared to the original excel data.
    :return:
    '''
    # --------------------- ER meningloma file read --------------------- #
    base_folder_path = '/home/soopil/Desktop/Dataset/brain_ewha'  # desktop setting
    train_path = os.path.join(base_folder_path, 'Train_Meningioma_20180508.xlsx')
    test_path = os.path.join(base_folder_path, 'Test_Meningioma_20180508.xlsx')
    xl = openpyxl.load_workbook(train_path, read_only=True)
    ws = xl['20171108_New_N4ITK corrected']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)
    # train_x = data_excel[1:, 4:102]
    feature_name_list = data_excel[0, 4:102]
    feature_T1 = []
    for e in feature_name_list:
        if 'T1C' in e:
            feature_T1.append(e)
            # print(e)
    # assert False
    # print(np.where('T1C' in feature_name_list))

    # --------------------- external file read --------------------- #
    base_folder_path = '/home/soopil/Desktop/Dataset/brain_ewha/external_validation'  # desktop setting
    external_path = os.path.join(base_folder_path, 'External validation set_Ewha_Meningioma.xlsx')
    xl = openpyxl.load_workbook(external_path, read_only=True)
    ws = xl['Sheet1']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)

    data_excel = np.array(data_excel)
    # print(data_excel[0,10:])
    feature_name_list_external = data_excel[0,10:]
    feature_T1_external = []
    for e in feature_name_list_external:
        if 'T1C' in e:
            feature_T1_external.append(e)
            # print(e)
    print(feature_T1)
    print(feature_T1_external)

    # for a,b in zip(feature_T1, feature_T1_external):
    #     print('{:30} | {:30}'.format(a,b) )

    l1, l2 = len(feature_T1), len(feature_T1_external)
    length = np.max([l1,l2])
    print(length)

    print('<{:30}> <{:30}>'.format('existing T1 features', 'features from external dataset'))
    print()
    for i in range(length):
        if i >= l1:
            a = ' - '
        else:
            a = feature_T1[i]
        if i >= l2:
            b = ' - '
        else:
            b = feature_T1_external[i]

        print('{:30} | {:30}'.format(a,b) )

def EWHA_CSV_reader():
    # --------------------- excel file read --------------------- #
    base_folder_path = "/home/soopil/Desktop/Dataset/EWHA_brain_tumor/0_Fwd_ MENINGIOMA 추가 자료 1_190711"
    external_path = os.path.join(base_folder_path,'Training set_Sinchon_Meningioma.xlsx')
    xl = openpyxl.load_workbook(external_path, read_only=True)
    ws = xl['20171108_New_N4ITK corrected']
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)

    data_excel = np.array(data_excel)
    subj_list_excel= data_excel[1:,0]
    label_list = np.squeeze(data_excel[1:, 5:5+1])
    sex_list = np.squeeze(data_excel[1:, 107:107+1])
    print(len(subj_list_excel),subj_list_excel)
    print(data_excel[0])
    print(label_list)
    print(sex_list)
    print('instance number of each class : ',len(np.where(label_list == 1)[0]),len(np.where(label_list == 0)[0]))

    # --------------------- Fractal dimention result file read --------------------- #
    result_dpath = '/home/soopil/Desktop/github/Ewha_brainMRI/fd_result/'
    # result_fname = 'SINCHON_FD_result_20190718.txt'
    # result_fname = 'SINCHON_FD_result_20190719_rescale.txt'
    result_fname = 'SINCHON_FD_result_20190723.txt' # original dataset
    result_fpath = os.path.join(result_dpath, result_fname)

    fd = open(result_fpath)
    lines = fd.readlines()
    print(lines)

    def FD(box_count):
        box_count = np.array(box_count).astype(np.float32)
        length = len(box_count)
        r = np.array([2 ** i for i in range(0, length)]).astype(np.float32)
        n_grad = np.gradient(np.log(box_count))
        r_grad = np.gradient(np.log(r))
        return (np.divide(n_grad, r_grad) * (-1))

    subj_list_fd, data_fd = [],[]

    for line in lines:

        if line == 'box counting fractal dimension.\n':
            continue
        split = line.split('/')
        subj = split[0]
        bx_count = split[2].split(',')
        lac = split[3].split(',')
        lac = list(np.array(lac).astype(np.float32))
        fdim = list(FD(bx_count))
        # print(fdim+lac)
        assert len(fdim) == 10 and len(lac) == 9
        subj_list_fd.append(subj)
        data_fd.append(fdim+lac)
    fd.close()

    # need certain normalization or other processes
    # ===================== normalize FD features ===================== #
    is_norm = False
    if is_norm:
        # normalize_col()
        data_fd = np.array(data_fd)
        maxi = np.amax(data_fd[:,:10])
        mini = np.amin(data_fd[:,:10])
        data_fd[:,:10] = (data_fd[:,:10] - mini) / maxi
        print(maxi, mini)
        # for e in data_fd:
        #     print(e)
        # assert False

    # ===================== csv file read ===================== #
    data_f, label_f  = [],[]
    subj_list_csv = []
    subj_list_no_csv = []
    csv_dir_path = os.path.join(base_folder_path, 'CSV')
    csv_file_list = os.listdir(csv_dir_path)
    print('total csv file number : ',len(csv_file_list))
    print(csv_file_list)

    for i, subj in enumerate(subj_list_excel):
        label, sex = label_list[i], sex_list[i]

        if sex == 'M':
            sex = [0.]
        else:
            sex = [1.]

        # print(subj, label, sex)
        T1C = str(subj)+'_T1C.csv'
        T2 = str(subj)+'_T2.csv'

        if (T1C in csv_file_list) and (T2 in csv_file_list):
            '''
            only use files with mask, T1C csv, T2C csv
            need to normalize features from T1C and T2 here
            '''
            # T1C csv file read
            csv_path = os.path.join(csv_dir_path, T1C)
            contents = read_csv(csv_path)
            # from 38 line Voxel volume
            feature_name = contents[:, 0]
            i_start = np.where(feature_name == 'original')[0][0]
            i_end = np.where(feature_name == 'original')[0][-1]
            length = i_end - i_start
            assert length == 106
            features_T1C = contents[i_start:i_start + length:, 3].astype(np.float32)

            if check_nan(features_T1C):
                print('detect nan number in T1C features : ',subj)

            # T2 csv file read
            csv_path = os.path.join(csv_dir_path, T2)
            contents = read_csv(csv_path)

            # format exception => skip first
            if subj in [5596598, 8453071]:
                subj_list_no_csv.append(subj)
                continue
                f = open(csv_path, 'r', encoding='utf-8')
                rdr = csv.reader(f)
                assert False

            # from 38 line Voxel volume
            feature_name = contents[:, 0]
            i_start = np.where(feature_name == 'original')[0][0]
            i_end = np.where(feature_name == 'original')[0][-1]
            length = i_end - i_start
            assert length == 106
            features_T2 = contents[i_start:i_start + length:, 3].astype(np.float32)

            if check_nan(features_T2):
                print('detect nan number in T2 features : ', subj)

            # sex features label
            is_FD_set = False
            features = sex + list(features_T1C) + list(features_T2)
            # features = []
            if is_FD_set:
                index_fd = subj_list_fd.index(str(subj))
                fd = list(data_fd[index_fd][:10]) #10 is fd and 9 is LAC
                features = features + fd

            # print(features)
            subj_list_csv.append(subj)
            data_f.append(features)
            label_f.append(label)
        else:
            subj_list_no_csv.append(subj)

    print('subjects with CSV files : ', len(subj_list_csv), subj_list_csv)
    print('subjects with no CSV files : ', len(subj_list_no_csv), subj_list_no_csv)
    return subj_list_csv, data_f, label_f

def read_xl(xl_path, sheet_name):
    xl = openpyxl.load_workbook(xl_path, read_only=True)
    ws = xl[sheet_name]
    data_excel = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        data_excel.append(line)
    data_excel = np.array(data_excel)
    print(data_excel.shape)
    subj_list = data_excel[1:, 0]
    label_list = data_excel[1:, 2]
    print(label_list)
    assert len(subj_list) == len(label_list)
    return subj_list, label_list

if __name__ == '__main__':
    # EWHA_excel_datareader()
    # EWHA_external_datareader()
    # check_features()
    subj_list, data, label = EWHA_CSV_reader()

    # for e in subj_list:
    #     print(e)
    # assert False

    low, high = [], []
    subj_low, subj_high = [], []
    for i in range(len(label)):
        if label[i] == 0:
            subj_low.append(subj_list[i])
            low.append(data[i])
        elif label[i] == 1:
            subj_high.append(subj_list[i])
            high.append(data[i])

    tTestResult = stats.ttest_ind(low, high)
    # print(tTestResult[0])
    # print(list(tTestResult[1]))

    print(subj_high[:5])
    print(subj_low[:5])

    pv = tTestResult[1]
    print('t test feature count on pvalue under the 0.05 : ',np.shape(np.where(pv < 0.05)))
    print(pv)
    print(np.where(pv < 0.05))
    # for i, v in enumerate(pv):
    #     print(v)
    print(high)

    # subj_list, data, label = SINCHON_NN_reader()
    # print(label)
    print('final class count : ',len(label), np.sum(label), len(label) - np.sum(label))
    # whole_set = split_data_by_fold(data, label, 5)
